import time
import openpyxl

from models.ctr import ClassToRegister


def is_header_row_found(attribute_indexes):
    for key in attribute_indexes:
        if attribute_indexes[key] != -1:
            return True
    return False


column_name_to_attribute_name = {
    "Kỳ": "term_id",
    "Mã_lớp": "class_id",
    "Mã_lớp_kèm": "second_class_id",
    "Mã_HP": "subject_id",
    "Tên_HP": "subject_name",
    "Buổi_số": "learn_day_number",
    "Thứ": "learn_at_day_of_week",
    "Phòng": "learn_room",
    "Thời_gian": "learn_time",
    "Tuần": "learn_week",
    "Loại_lớp": "class_type",
    "Ghi_chú": "describe",
}


class Parser:

    def __init__(self, id=str(time.time_ns())):
        self.id = id
        self.attribute_indexes = {
            "term_id": -1,
            "class_id": -1,
            "second_class_id": -1,
            "subject_id": -1,
            "subject_name": -1,
            "learn_day_number": -1,
            "learn_at_day_of_week": -1,
            "learn_room": -1,
            "learn_time": -1,
            "learn_week": -1,
            "class_type": -1,
            "describe": -1,
        }

    def parse(self, file):
        work_book = openpyxl.load_workbook(file)
        work_sheet = work_book.active
        max_column = work_sheet.max_column
        max_row = work_sheet.max_row
        current_row = 0
        class_list = []
        term_ids = set()

        print(f" [*] {self.id} max_column = {max_column}")
        print(f" [*] {self.id} max_row = {max_row}")

        iterator = work_sheet.iter_rows(min_row=0, values_only=True)

        for row in iterator:
            current_row += 1
            i = 0
            for cell in row:
                prop_name = column_name_to_attribute_name.get(str(cell), None)
                if prop_name:
                    self.attribute_indexes[prop_name] = i
                i = i + 1
            if is_header_row_found(self.attribute_indexes):
                # found header of data rows
                break

        for row in iterator:
            construct_opts = {}
            for key in self.attribute_indexes:
                construct_opts[key] = row[self.attribute_indexes[key]]
            c = ClassToRegister(**construct_opts)
            term_ids.add(c.term_id)
            class_list.append(c)

        print(f" [*] {self.id} class_to_register_count = {len(class_list)}")
        print(f" [*] {self.id} term_ids = {list(term_ids)}")
        print(f" [*] {self.id} sample_class_to_register = {class_list[0]}")

        return class_list
